"""
SPS API Views
REST API ViewSets with proper permissions and filtering
"""
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser
from django.shortcuts import get_object_or_404
from django.db.models import Count, Avg, Q, Sum
from django.db import transaction,models
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.exceptions import PermissionDenied, NotFound
from django.db.models import Count, Max
import openpyxl
import tempfile
import os
import random
from openpyxl import load_workbook
import csv

from .models import (
    Subject, Topic, Resource, Quiz, Question,
    Choice, QuizAttempt, StudentResponse
)
from .serializers import (
    SubjectSerializer, TopicSerializer, ResourceSerializer,
    QuizSerializer, QuizPublicSerializer, QuestionSerializer,
    QuizAttemptSerializer, QuizSubmissionSerializer,
    LeaderboardEntrySerializer, QuizStatsSerializer
)
from .permissions import IsTeacher, IsOwner
from .pagination import StandardResultsSetPagination


class SubjectViewSet(viewsets.ModelViewSet):
    """Subject CRUD operations"""
    serializer_class = SubjectSerializer
    permission_classes = [IsAuthenticated, IsTeacher]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['subject_name', 'description']
    ordering_fields = ['created_at', 'subject_name']
    ordering = ['-created_at']
    lookup_field = 'subject_uuid'
    lookup_value_regex = '[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}'

    def get_queryset(self):
        return Subject.objects.filter(
            teacher=self.request.user
        ).annotate(
            topics_count=Count('topics'),
            quizzes_count=Count('topics__quizzes')
        )

    def perform_create(self, serializer):
        serializer.save(teacher=self.request.user)

    @action(detail=True, methods=['get'])
    def stats(self, request, *args, **kwargs):
        """Get detailed statistics for a subject"""
        subject = self.get_object()
        topics = Topic.objects.filter(subject=subject, teacher=request.user)
        quizzes = Quiz.objects.filter(topic__in=topics)

        total_attempts = QuizAttempt.objects.filter(
            quiz__in=quizzes, is_completed=True
        ).count()

        passed_count = sum(
            1 for a in QuizAttempt.objects.filter(quiz__in=quizzes, is_completed=True)
            if a.percentage >= 70
        )

        return Response({
            'topics_count': topics.count(),
            'quizzes_count': quizzes.count(),
            'total_attempts': total_attempts,
            'passed_count': passed_count,
            'failed_count': total_attempts - passed_count
        })




@api_view(['GET'])
@permission_classes([AllowAny])
def quiz_join_by_code(request, session_code):
    """Student session code orqali quizga qo'shiladi"""
    quiz = get_object_or_404(
        Quiz,
        session_code=session_code,
        is_active=True,
        session_ended=False
    )

    serializer = QuizPublicSerializer(quiz, context={'request': request})
    return Response(serializer.data)



class TopicViewSet(viewsets.ModelViewSet):
    """Topic CRUD operations"""
    serializer_class = TopicSerializer
    permission_classes = [IsAuthenticated, IsTeacher]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['topic_name', 'description']
    ordering_fields = ['order', 'created_at']
    ordering = ['order', 'created_at']
    lookup_field = 'topic_uuid'
    lookup_value_regex = '[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}'

    def get_queryset(self):
        queryset = Topic.objects.filter(
            teacher=self.request.user
        ).select_related('subject').annotate(
            resources_count=Count('resources'),
            quizzes_count=Count('quizzes')
        )
        subject_uuid = self.request.query_params.get('subject_uuid')
        if subject_uuid:
            queryset = queryset.filter(subject__subject_uuid=subject_uuid)
        return queryset

    def perform_create(self, serializer):
        """Create topic"""
        subject_uuid = serializer.validated_data.get('subject_uuid')

        try:
            subject = Subject.objects.get(
                subject_uuid=subject_uuid,
                teacher=self.request.user
            )
        except Subject.DoesNotExist:
            raise NotFound("Subject not found or you don't have permission")

        last_order = Topic.objects.filter(subject=subject).aggregate(
            max_order=Max('order')
        )['max_order'] or 0

        serializer.save(teacher=self.request.user, subject=subject, order=last_order + 1)

    def perform_update(self, serializer):
        """Update topic - only name and description"""
        topic = self.get_object()

        if topic.teacher != self.request.user:
            raise PermissionDenied("You can only update your own topics")

        # Ignore subject_uuid if provided - subject can't be changed
        if 'subject_uuid' in serializer.validated_data:
            serializer.validated_data.pop('subject_uuid')

        serializer.save()

    def destroy(self, request, *args, **kwargs):
        """Delete topic with force option"""
        instance = self.get_object()

        if instance.teacher != request.user:
            raise PermissionDenied("You can only delete your own topics")

        force = request.query_params.get('force', 'false').lower() == 'true'

        has_resources = instance.resources.exists()
        has_quizzes = instance.quizzes.exists()

        if (has_resources or has_quizzes) and not force:
            resources_count = instance.resources.count()
            quizzes_count = instance.quizzes.count()

            items = []
            if has_resources:
                items.append(f"{resources_count} resource(s)")
            if has_quizzes:
                items.append(f"{quizzes_count} quiz(zes)")

            return Response({
                'error': f"This topic contains {' and '.join(items)}.",
                'message': 'Please delete or move them to another topic before deleting, or use force=true to delete everything.',
                'has_resources': has_resources,
                'has_quizzes': has_quizzes,
                'resources_count': resources_count,
                'quizzes_count': quizzes_count,
                'can_force_delete': True
            }, status=status.HTTP_400_BAD_REQUEST)

        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'])
    def reorder(self, request):
        """Reorder topics"""
        orders = request.data.get('orders', [])

        if not orders:
            return Response(
                {'error': 'orders list is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            for item in orders:
                topic_uuid = item.get('topic_uuid')
                order = item.get('order')

                if not topic_uuid or order is None:
                    continue

                Topic.objects.filter(
                    topic_uuid=topic_uuid,
                    teacher=request.user
                ).update(order=order)

        return Response({'status': 'success'})

class ResourceViewSet(viewsets.ModelViewSet):
    """Resource CRUD operations"""
    serializer_class = ResourceSerializer
    permission_classes = [IsAuthenticated, IsTeacher]
    pagination_class = StandardResultsSetPagination
    parser_classes = [MultiPartParser, FormParser]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    ordering_fields = ['order', 'created_at']
    ordering = ['order', 'created_at']
    lookup_field = 'resource_uuid'
    lookup_value_regex = '[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}'  # ← ADD THIS LINE

    def get_queryset(self):
        queryset = Resource.objects.filter(teacher=self.request.user)

        # "topic_uuid" yoki "topic" ikkalasini ham qabul qiladi
        topic_uuid = (
                self.request.query_params.get('topic_uuid') or
                self.request.query_params.get('topic')
        )

        if topic_uuid:
            # Oxiridagi slash va bo'shliqlarni tozalash
            topic_uuid = topic_uuid.strip('/ ')
            queryset = queryset.filter(topic__topic_uuid=topic_uuid)

        return queryset

    def perform_create(self, serializer):
        print(serializer)
        topic = serializer.validated_data['topic']
        if topic.teacher != self.request.user:
            raise PermissionError("You can only add resources to your own topics")

        # Get last order
        last_order = Resource.objects.filter(topic=topic).aggregate(
            max_order=models.Max('order')
        )['max_order'] or 0

        # Detect file type
        file = serializer.validated_data['file']
        filename = file.name.lower()

        if filename.endswith('.pdf'):
            file_type = 'pdf'
        elif filename.endswith(('.ppt', '.pptx')):
            file_type = 'pptx'
        elif filename.endswith(('.doc', '.docx')):
            file_type = 'docx'
        else:
            file_type = 'other'

        serializer.save(
            teacher=self.request.user,
            order=last_order + 1,
            file_type=file_type,
            file_size=file.size
        )

    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        files = request.FILES.getlist('files')
        subject_uuid = request.data.get('subject_uuid')

        if not files:
            return Response({'error': 'Kamida 1 ta fayl yuklang'}, status=400)

        if len(files) > 10:
            return Response({'error': 'Maksimum 10 ta fayl'}, status=400)

        if not subject_uuid:
            return Response({'error': 'subject_uuid required'}, status=400)

        try:
            subject = Subject.objects.get(subject_uuid=subject_uuid, teacher=request.user)
        except Subject.DoesNotExist:
            raise NotFound("Subject topilmadi")

        created = []
        errors = []

        # Boshlang'ich order bir marta olamiz
        base_order = Topic.objects.filter(subject=subject).aggregate(
            max_order=Max('order')
        )['max_order'] or 0

        for idx, file in enumerate(files):
            try:
                filename = file.name
                name_without_ext = filename.rsplit('.', 1)[0]

                filename_lower = filename.lower()
                if filename_lower.endswith('.pdf'):
                    file_type = 'pdf'
                elif filename_lower.endswith(('.ppt', '.pptx')):
                    file_type = 'pptx'
                elif filename_lower.endswith(('.doc', '.docx')):
                    file_type = 'docx'
                else:
                    file_type = 'other'

                # Har bir fayl uchun alohida topic
                topic = Topic.objects.create(
                    subject=subject,
                    teacher=request.user,
                    topic_name=name_without_ext,
                    order=base_order + idx + 1  # ← idx bilan incrementing
                )

                # Faqat SHU topic uchun SHU fayl
                resource = Resource.objects.create(
                    topic=topic,
                    teacher=request.user,
                    title=name_without_ext,
                    file=file,
                    file_type=file_type,
                    file_size=file.size,
                    order=1
                )

                created.append({
                    'topic_uuid': str(topic.topic_uuid),
                    'topic_name': topic.topic_name,
                    'resource_uuid': str(resource.resource_uuid),
                    'file_type': file_type,
                })

            except Exception as e:
                errors.append({'filename': file.name, 'error': str(e)})

        return Response({
            'created': created,
            'errors': errors,
            'total_created': len(created),
        }, status=201)

    @action(detail=False, methods=['post'])
    def reorder(self, request):
        """Reorder resources"""
        orders = request.data.get('orders', [])

        with transaction.atomic():
            for item in orders:
                Resource.objects.filter(
                    id=item['id'],
                    teacher=request.user
                ).update(order=item['order'])

        return Response({'status': 'success'})



class QuizViewSet(viewsets.ModelViewSet):
    """Quiz CRUD operations (Teacher)"""
    serializer_class = QuizSerializer
    permission_classes = [IsAuthenticated, IsTeacher]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['title', 'description']
    ordering_fields = ['created_at', 'title']
    ordering = ['-created_at']
    lookup_field = 'quiz_uuid'
    lookup_value_regex = '[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}'

    def get_queryset(self):
        user_topics = Topic.objects.filter(teacher=self.request.user)
        queryset = Quiz.objects.filter(
            topic__in=user_topics
        ).select_related('topic', 'topic__subject').annotate(
            attempt_count=Count('attempts', filter=Q(attempts__is_completed=True)),
            avg_score=Avg('attempts__score', filter=Q(attempts__is_completed=True)),
            questions_count=Count('questions')
        )

        topic_uuid = self.request.query_params.get('topic_uuid')
        if topic_uuid:
            queryset = queryset.filter(topic__topic_uuid=topic_uuid)

        return queryset

    def perform_create(self, serializer):
        """Create quiz with topic_uuid"""
        topic_uuid = serializer.validated_data.get('topic_uuid')

        if not topic_uuid:
            raise ValidationError({'topic_uuid': 'This field is required'})

        try:
            topic = Topic.objects.get(
                topic_uuid=topic_uuid,
                teacher=self.request.user
            )
        except Topic.DoesNotExist:
            raise NotFound("Topic not found or you don't have permission")

        quiz = serializer.save(teacher=self.request.user, topic=topic)
        quiz.generate_session_code()

    def perform_update(self, serializer):
        """Update quiz - only basic fields"""
        quiz = self.get_object()

        if quiz.topic.teacher != self.request.user:
            raise PermissionDenied("You can only update your own quizzes")

        # Remove topic_uuid if provided (topic can't be changed)
        if 'topic_uuid' in serializer.validated_data:
            serializer.validated_data.pop('topic_uuid')

        serializer.save()

    def destroy(self, request, *args, **kwargs):
        """Delete quiz"""
        instance = self.get_object()

        if instance.topic.teacher != request.user:
            raise PermissionDenied("You can only delete your own quizzes")

        # Check if quiz has attempts
        has_attempts = instance.attempts.exists()
        force = request.query_params.get('force', 'false').lower() == 'true'

        if has_attempts and not force:
            attempts_count = instance.attempts.count()

            return Response({
                'error': f"This quiz has {attempts_count} attempt(s).",
                'message': 'Deleting will remove all student results. Use force=true to confirm deletion.',
                'attempts_count': attempts_count,
                'can_force_delete': True
            }, status=status.HTTP_400_BAD_REQUEST)

        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'], url_path='regenerate-code')
    def regenerate_code(self, request, *args, **kwargs):
        """Regenerate session code"""
        quiz = self.get_object()

        if quiz.topic.teacher != request.user:
            raise PermissionDenied("You can only regenerate code for your own quizzes")

        quiz.end_session()
        new_code = quiz.generate_session_code()
        return Response({'session_code': new_code})

    @action(detail=True, methods=['post'], url_path='end-session')
    def end_session(self, request, *args, **kwargs):
        """End quiz session"""
        quiz = self.get_object()

        if quiz.topic.teacher != request.user:
            raise PermissionDenied("You can only end session for your own quizzes")

        quiz.end_session()
        return Response({'status': 'Session ended'})

    @action(detail=True, methods=['get'])
    def results(self, request, *args, **kwargs):
        """Get quiz results and statistics"""
        quiz = self.get_object()

        if quiz.topic.teacher != request.user:
            raise PermissionDenied("You can only view results for your own quizzes")

        attempts = QuizAttempt.objects.filter(
            quiz=quiz, is_completed=True
        ).order_by('-score', 'completed_at')

        max_score = sum(q.points for q in quiz.questions.all())
        total_students = attempts.count()
        passed_count = sum(1 for a in attempts if a.percentage >= 70)
        avg_score = attempts.aggregate(avg=Avg('score'))['avg'] or 0

        stats = QuizStatsSerializer({
            'total_attempts': total_students,
            'average_score': round(avg_score, 1),
            'passed_count': passed_count,
            'failed_count': total_students - passed_count,
            'max_possible_score': max_score
        })

        serializer = QuizAttemptSerializer(attempts, many=True)

        return Response({
            'quiz': QuizSerializer(quiz).data,
            'stats': stats.data,
            'attempts': serializer.data
        })

    @action(detail=True, methods=['get'], url_path='export-excel')
    def export_excel(self, request, *args, **kwargs):
        """Export quiz results to Excel"""
        quiz = self.get_object()

        if quiz.topic.teacher != request.user:
            raise PermissionDenied("You can only export results for your own quizzes")

        attempts = QuizAttempt.objects.filter(
            quiz=quiz, is_completed=True
        ).order_by('-score')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"

        headers = ["#", "Student Name", "Score", "Max Score", "Percentage", "Status", "Completed At"]
        ws.append(headers)

        for idx, attempt in enumerate(attempts, 1):
            pct = attempt.percentage
            status_text = "Passed ✓" if pct >= 70 else "Failed ✗"
            ws.append([
                idx,
                attempt.student_name,
                attempt.score,
                attempt.max_score,
                f"{pct}%",
                status_text,
                attempt.completed_at.strftime("%Y-%m-%d %H:%M") if attempt.completed_at else ""
            ])

        from django.http import HttpResponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{quiz.title}_results.xlsx"'
        wb.save(response)
        return response

# Public Quiz Views (No Authentication Required)

@api_view(['GET'])
@permission_classes([AllowAny])
def quiz_public_detail(request, quiz_uuid):
    """Get quiz details for students (public)"""
    quiz = get_object_or_404(Quiz, quiz_uuid=quiz_uuid, is_active=True)

    if quiz.session_ended:
        return Response(
            {'error': 'Quiz session has ended'},
            status=status.HTTP_410_GONE
        )

    serializer = QuizPublicSerializer(quiz, context={'request': request})
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([AllowAny])
def quiz_submit(request, quiz_uuid):
    """Submit quiz answers (public)"""
    quiz = get_object_or_404(Quiz, quiz_uuid=quiz_uuid, is_active=True)

    if quiz.session_ended:
        return Response(
            {'error': 'Quiz session has ended'},
            status=status.HTTP_410_GONE
        )

    serializer = QuizSubmissionSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    student_name = serializer.validated_data['student_name']
    answers = serializer.validated_data['answers']

    with transaction.atomic():
        # Create attempt
        attempt = QuizAttempt.objects.create(
            quiz=quiz,
            student_name=student_name,
            started_at=timezone.now()
        )

        # Process answers
        questions = quiz.questions.prefetch_related('choices').all()
        max_score = sum(q.points for q in questions)
        total_score = 0.0

        for question in questions:
            # Get student's selected choice IDs for this question
            answer_data = next(
                (a for a in answers if a.get('question_id') == question.id),
                None
            )
            selected_ids = answer_data.get('choice_ids', []) if answer_data else []

            # Validate selected choices
            valid_choices = list(
                question.choices.filter(id__in=selected_ids).values_list('id', flat=True)
            )

            # Check if correct
            correct_ids = sorted(
                question.choices.filter(is_correct=True).values_list('id', flat=True)
            )
            is_correct = sorted(valid_choices) == correct_ids and len(correct_ids) > 0
            earned = question.points if is_correct else 0

            # Save response
            response = StudentResponse.objects.create(
                attempt=attempt,
                question=question,
                is_correct=is_correct,
                earned_points=earned
            )
            if valid_choices:
                response.selected_choices.set(valid_choices)

            total_score += earned

        # Update attempt
        attempt.score = total_score
        attempt.max_score = max_score
        attempt.is_completed = True
        attempt.completed_at = timezone.now()
        attempt.save()

    # Return result
    serializer = QuizAttemptSerializer(attempt)
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([AllowAny])
def quiz_result(request, attempt_uuid):
    """Get quiz attempt result (public)"""
    attempt = get_object_or_404(
        QuizAttempt,
        attempt_uuid=attempt_uuid,
        is_completed=True
    )

    serializer = QuizAttemptSerializer(attempt)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([AllowAny])
def quiz_leaderboard(request, quiz_uuid):
    """Get live leaderboard (public)"""
    quiz = get_object_or_404(Quiz, quiz_uuid=quiz_uuid)

    attempts = QuizAttempt.objects.filter(
        quiz=quiz,
        is_completed=True
    ).order_by('-score', 'completed_at')[:30]

    leaderboard = []
    for idx, attempt in enumerate(attempts, 1):
        leaderboard.append({
            'rank': idx,
            'student_name': attempt.student_name,
            'score': attempt.score,
            'max_score': attempt.max_score,
            'percentage': attempt.percentage,
            'completed_at': attempt.completed_at
        })

    serializer = LeaderboardEntrySerializer(leaderboard, many=True)

    return Response({
        'total': attempts.count(),
        'session_ended': quiz.session_ended,
        'leaderboard': serializer.data
    })


class QuestionViewSet(viewsets.ModelViewSet):
    """Question CRUD operations (Teacher) - Nested under Quiz"""
    serializer_class = QuestionSerializer
    permission_classes = [IsAuthenticated, IsTeacher]
    lookup_field = 'question_uuid'
    lookup_value_regex = '[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}'

    def get_queryset(self):
        """Get questions for a specific quiz"""
        quiz_uuid = self.kwargs.get('quiz_uuid')

        try:
            quiz = Quiz.objects.get(quiz_uuid=quiz_uuid)

            # Check if teacher owns this quiz
            if quiz.topic.teacher != self.request.user:
                return Question.objects.none()

            return Question.objects.filter(quiz=quiz).prefetch_related('choices')
        except Quiz.DoesNotExist:
            return Question.objects.none()

    def get_quiz(self):
        """Helper to get quiz and check permission"""
        quiz_uuid = self.kwargs.get('quiz_uuid')

        try:
            quiz = Quiz.objects.select_related('topic__teacher').get(quiz_uuid=quiz_uuid)

            if quiz.topic.teacher != self.request.user:
                raise PermissionDenied("You can only manage questions in your own quizzes")

            return quiz
        except Quiz.DoesNotExist:
            raise NotFound("Quiz not found")

    def perform_create(self, serializer):
        """Create question in quiz"""
        quiz = self.get_quiz()
        serializer.save(quiz=quiz)

    def perform_update(self, serializer):
        """Update question"""
        quiz = self.get_quiz()
        question = self.get_object()

        if question.quiz != quiz:
            raise PermissionDenied("Question does not belong to this quiz")

        serializer.save()

    def destroy(self, request, *args, **kwargs):
        """Delete question"""
        quiz = self.get_quiz()
        instance = self.get_object()

        if instance.quiz != quiz:
            raise PermissionDenied("Question does not belong to this quiz")

        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsTeacher])
def quiz_preview_excel(request):
    """Excel/CSV yuklaydi, columnlar va 5ta preview row qaytaradi."""
    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'file required'}, status=400)

    filename = file.name.lower()
    rows = []

    try:
        if filename.endswith('.csv'):
            import io, csv
            content = file.read().decode('utf-8-sig')
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)
        elif filename.endswith(('.xlsx', '.xls')):
            from openpyxl import load_workbook
            wb = load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else '' for cell in row])
            wb.close()
        else:
            return Response({'error': 'Faqat .xlsx, .xls yoki .csv'}, status=400)
    except Exception as e:
        return Response({'error': f'Fayl o\'qishda xato: {str(e)}'}, status=400)

    if not rows:
        return Response({'error': 'Fayl bo\'sh'}, status=400)

    headers = rows[0]
    preview_rows = rows[1:6]

    columns = []
    for idx, h in enumerate(headers):
        columns.append({
            'index': idx,
            'name': str(h),
            'preview': [row[idx] if idx < len(row) else '' for row in preview_rows]
        })

    return Response({
        'columns': columns,
        'total_rows': len(rows) - 1,
        'preview_count': len(preview_rows),
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsTeacher])
def quiz_import_excel(request):
    """Excel/CSV + column mapping → Quiz + Question + Choice yaratadi."""
    import io, csv, json, random
    from openpyxl import load_workbook

    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'file required'}, status=400)

    topic_uuid      = request.data.get('topic_uuid')
    title           = request.data.get('title', '').strip()
    description     = request.data.get('description', '')
    time_limit      = int(request.data.get('time_limit', 15))
    question_col    = request.data.get('question_col')
    correct_col     = request.data.get('correct_col')
    choice_cols_raw = request.data.get('choice_cols', '[]')
    points_per_q    = int(request.data.get('points_per_question', 5))

    errors = {}
    if not topic_uuid:       errors['topic_uuid']     = 'required'
    if not title:            errors['title']           = 'required'
    if question_col is None: errors['question_col']   = 'required'
    if correct_col is None:  errors['correct_col']    = 'required'
    if errors:
        return Response(errors, status=400)

    try:
        question_col = int(question_col)
        correct_col  = int(correct_col)
        choice_cols  = json.loads(choice_cols_raw) if isinstance(choice_cols_raw, str) else choice_cols_raw
        choice_cols  = [int(c) for c in choice_cols]
    except (ValueError, json.JSONDecodeError):
        return Response({'error': 'column indexlar integer bo\'lishi kerak'}, status=400)

    if not choice_cols:
        return Response({'error': 'Kamida 1ta qo\'shimcha variant ustuni tanlang'}, status=400)

    try:
        topic = Topic.objects.get(topic_uuid=topic_uuid, teacher=request.user)
    except Topic.DoesNotExist:
        return Response({'error': 'Topic topilmadi'}, status=404)

    # Faylni o'qish
    filename = file.name.lower()
    rows = []
    try:
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)
        elif filename.endswith(('.xlsx', '.xls')):
            wb = load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else '' for cell in row])
            wb.close()
        else:
            return Response({'error': 'Faqat .xlsx, .xls yoki .csv'}, status=400)
    except Exception as e:
        return Response({'error': f'Fayl o\'qishda xato: {str(e)}'}, status=400)

    data_rows = rows[1:]
    if not data_rows:
        return Response({'error': 'Faylda ma\'lumot yo\'q'}, status=400)

    with transaction.atomic():
        quiz = Quiz.objects.create(
            topic=topic,
            teacher=request.user,
            title=title,
            description=description,
            time_limit=time_limit,
        )
        quiz.generate_session_code()

        created_questions = 0
        skipped = 0

        for order_idx, row in enumerate(data_rows):
            def get_col(idx):
                try:
                    val = row[idx]
                    return str(val).strip() if val and str(val).strip() not in ('', 'None') else ''
                except IndexError:
                    return ''

            question_text  = get_col(question_col)
            correct_answer = get_col(correct_col)

            if not question_text or not correct_answer:
                skipped += 1
                continue

            wrong_choices = []
            for ci in choice_cols:
                val = get_col(ci)
                if val:
                    wrong_choices.append(val)

            question = Question.objects.create(
                quiz=quiz,
                text=question_text,
                question_type='single',
                points=points_per_q,
                order=order_idx,
            )

            all_choices = [(correct_answer, True)] + [(w, False) for w in wrong_choices]
            random.shuffle(all_choices)

            for choice_order, (text, is_correct) in enumerate(all_choices):
                Choice.objects.create(
                    question=question,
                    text=text,
                    is_correct=is_correct,
                    order=choice_order,
                )

            created_questions += 1

    return Response({
        'quiz_uuid':         str(quiz.quiz_uuid),
        'session_code':      quiz.session_code,
        'title':             quiz.title,
        'topic_name':        topic.topic_name,
        'created_questions': created_questions,
        'skipped_rows':      skipped,
    }, status=201)
